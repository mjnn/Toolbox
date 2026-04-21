import csv
import io
import json
import re
from typing import List, Literal

from email_validator import EmailNotValidError, validate_email
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlalchemy import func, or_
from sqlmodel import Session, select

from app.database import get_session
from app.models import (
    UserToolPermission,
    Tool,
    ToolRelease,
    User,
    PermissionStatus,
    Notification,
    ToolOwner,
    Role,
    UserRole,
    APIAccessLog,
    Feedback,
    FeedbackCategory,
    ToolAnnouncement,
)
from app.schemas import (
    AdminUserImportIssue,
    AdminUserImportResponse,
    PermissionWithDetails,
    PermissionUpdate,
    SuccessResponse,
    ToolInDB,
    ToolOwnerWithUser,
    ToolLicenseUserRow,
    PaginatedToolLicenseUsers,
    ToolStatusUpdate,
    RoleAssignmentRequest,
    UserRolesResponse,
    APIAccessLogWithUser,
    PaginatedAPIAccessLogs,
    FeedbackWithUser,
    PaginatedFeedbackWithUser,
    FeedbackCountsResponse,
    UserInDB,
    ToolReleasePublish,
    ToolReleaseInDB,
    ToolAnnouncementCreate,
    ToolAnnouncementInDB,
    ToolAnnouncementUpdate,
    PaginatedToolAnnouncements,
)
from app.api.v1.users import get_current_active_user
from app.api.v1.auth import get_password_hash
from app.services.user_deletion import delete_user_and_related
from app.services.tool_behavior_catalog import resolve_behavior_label_from_tool
from datetime import datetime, timezone
from zoneinfo import ZoneInfo

router = APIRouter()

_TZ_CN = ZoneInfo("Asia/Shanghai")
_ANNOUNCEMENT_PRIORITIES = {"urgent", "notice", "reminder"}
_GLOBAL_ANNOUNCEMENT_TOOL_NAME = "mos-integration-toolbox"


def _format_ts_cst8(dt: datetime | None) -> str:
    """API / DB 中 naive UTC 时刻 → 东八区展示字符串（与前端 Intl Asia/Shanghai 一致）。"""
    if dt is None:
        return ""
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(_TZ_CN).strftime("%Y-%m-%d %H:%M:%S")


def get_role_by_name(db: Session, role_name: str) -> Role:
    role = db.exec(select(Role).where(Role.name == role_name)).first()
    if not role:
        raise HTTPException(status_code=404, detail=f"角色不存在：{role_name}")
    return role


def user_has_role(db: Session, user_id: int, role_name: str) -> bool:
    role = db.exec(select(Role).where(Role.name == role_name)).first()
    if not role:
        return False
    return db.exec(
        select(UserRole).where(UserRole.user_id == user_id, UserRole.role_id == role.id)
    ).first() is not None


def user_is_tool_owner(db: Session, user_id: int, tool_id: int) -> bool:
    return db.exec(
        select(ToolOwner).where(ToolOwner.user_id == user_id, ToolOwner.tool_id == tool_id)
    ).first() is not None


def ensure_admin(current_user: User):
    if not current_user.is_superuser:
        raise HTTPException(status_code=403, detail="仅管理员可执行该操作")


def _is_announcement_active(row: ToolAnnouncement, now: datetime) -> bool:
    if not row.is_enabled:
        return False
    if row.start_at and row.start_at > now:
        return False
    if row.end_at and row.end_at < now:
        return False
    return True


def _normalize_announcement_priority(raw: str | None) -> str:
    level = str(raw or "notice").strip().lower()
    if level not in _ANNOUNCEMENT_PRIORITIES:
        raise HTTPException(status_code=400, detail="公告优先级仅支持 urgent / notice / reminder")
    return level


def _priority_colors(priority: str) -> tuple[str, str]:
    if priority == "urgent":
        return "#ffffff", "#c62828"
    if priority == "reminder":
        return "#5f370e", "#fff4e5"
    return "#102a43", "#e8f4fd"


def _build_announcement_schema(row: ToolAnnouncement) -> ToolAnnouncementInDB:
    try:
        disable_feature_slugs = json.loads(row.disable_feature_slugs_json or "[]")
    except Exception:
        disable_feature_slugs = []
    if not isinstance(disable_feature_slugs, list):
        disable_feature_slugs = []
    return ToolAnnouncementInDB(
        id=row.id,
        tool_id=row.tool_id,
        title=row.title,
        content=row.content,
        is_enabled=row.is_enabled,
        start_at=row.start_at,
        end_at=row.end_at,
        visibility=row.visibility or "global",
        priority=row.priority or "notice",
        scroll_speed_seconds=int(row.scroll_speed_seconds or 45),
        font_family=row.font_family,
        font_size_px=int(row.font_size_px or 14),
        text_color=row.text_color,
        background_color=row.background_color,
        disable_feature_slugs=[str(v) for v in disable_feature_slugs if str(v).strip()],
        created_by=row.created_by,
        created_at=row.created_at,
        updated_at=row.updated_at,
    )


def _resolve_global_announcement_tool_id(db: Session) -> int:
    tool = db.exec(select(Tool).where(Tool.name == _GLOBAL_ANNOUNCEMENT_TOOL_NAME)).first()
    if not tool or not tool.id:
        raise HTTPException(status_code=500, detail="未找到全局公告绑定工具，请联系管理员")
    return int(tool.id)


def _normalize_excel_header(value: object) -> str:
    text = str(value or "").strip().lower()
    return re.sub(r"[\s_\-]+", "", text)


def _sanitize_username(raw: str) -> str:
    text = re.sub(r"[^a-zA-Z0-9_.-]+", "-", raw.strip().lower()).strip("-._")
    return text[:50] if text else "user"


def _pick_username(base: str, used_usernames: set[str], db: Session) -> str:
    base_candidate = _sanitize_username(base)
    candidate = base_candidate
    suffix = 1
    while True:
        if candidate not in used_usernames:
            exists = db.exec(select(User.id).where(User.username == candidate)).first()
            if not exists:
                used_usernames.add(candidate)
                return candidate
        suffix += 1
        clipped = base_candidate[:40] if len(base_candidate) > 40 else base_candidate
        candidate = f"{clipped}-{suffix}"


def _extract_email(raw: object) -> str:
    email = str(raw or "").strip().lower()
    if not email:
        raise ValueError("邮箱为空")
    try:
        normalized = validate_email(email, check_deliverability=False)
    except EmailNotValidError as exc:
        raise ValueError(f"邮箱格式非法：{exc}") from exc
    return normalized.normalized


def ensure_permission_reviewer(db: Session, current_user: User, tool_id: int):
    if current_user.is_superuser:
        return
    if user_is_tool_owner(db, current_user.id, tool_id):
        return
    raise HTTPException(status_code=403, detail="仅工具负责人可审核该权限")


def build_access_log_item(db: Session, access_log: APIAccessLog) -> APIAccessLogWithUser:
    user = db.get(User, access_log.user_id) if access_log.user_id else None
    data = access_log.dict()
    bl = data.get("behavior_label")
    if not bl and access_log.tool_id and access_log.feature_name:
        tool = db.get(Tool, access_log.tool_id)
        data["behavior_label"] = resolve_behavior_label_from_tool(tool, access_log.feature_name)
    return APIAccessLogWithUser(**data, user=user)


def build_feedback_with_user(db: Session, fb: Feedback) -> FeedbackWithUser:
    user = db.get(User, fb.user_id)
    if not user:
        raise HTTPException(status_code=500, detail="反馈数据异常：用户不存在")
    u = UserInDB.model_validate(user)
    return FeedbackWithUser(
        id=fb.id,
        user_id=fb.user_id,
        tool_id=fb.tool_id,
        category=fb.category,
        title=fb.title,
        content=fb.content,
        created_at=fb.created_at,
        user=u,
    )


@router.get("/announcements/global", response_model=PaginatedToolAnnouncements)
async def list_global_announcements(
    skip: int = 0,
    limit: int = 20,
    only_active: bool = False,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_session),
):
    ensure_admin(current_user)
    limit = min(max(limit, 1), 200)
    rows = db.exec(
        select(ToolAnnouncement)
        .where(ToolAnnouncement.visibility == "global")
        .order_by(ToolAnnouncement.created_at.desc())
    ).all()
    if only_active:
        now = datetime.utcnow()
        rows = [row for row in rows if _is_announcement_active(row, now)]
    total = len(rows)
    page_items = rows[skip : skip + limit]
    return PaginatedToolAnnouncements(
        total=total,
        items=[_build_announcement_schema(row) for row in page_items],
    )


@router.post("/announcements/global", response_model=ToolAnnouncementInDB)
async def create_global_announcement(
    body: ToolAnnouncementCreate,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_session),
):
    ensure_admin(current_user)
    tool_id = _resolve_global_announcement_tool_id(db)
    if body.end_at and body.start_at and body.end_at < body.start_at:
        raise HTTPException(status_code=400, detail="结束时间不能早于开始时间")
    now = datetime.utcnow()
    priority = _normalize_announcement_priority(body.priority)
    text_color, background_color = _priority_colors(priority)
    row = ToolAnnouncement(
        tool_id=tool_id,
        visibility="global",
        priority=priority,
        title=body.title.strip(),
        content=body.content.strip(),
        is_enabled=body.is_enabled,
        start_at=body.start_at,
        end_at=body.end_at,
        scroll_speed_seconds=body.scroll_speed_seconds,
        font_family=(body.font_family or "").strip() or None,
        font_size_px=body.font_size_px,
        text_color=text_color,
        background_color=background_color,
        disable_feature_slugs_json="[]",
        created_by=current_user.id,
        created_at=now,
        updated_at=now,
    )
    db.add(row)
    db.commit()
    db.refresh(row)
    if row.is_enabled:
        users = db.exec(
            select(User).where(
                User.is_active == True,  # noqa: E712
                User.is_approved == True,  # noqa: E712
            )
        ).all()
        for u in users:
            db.add(
                Notification(
                    user_id=u.id,
                    title=f"系统公告：{row.title}",
                    message=row.content,
                    notification_type="system",
                    related_id=row.id,
                )
            )
        db.commit()
    return _build_announcement_schema(row)


@router.patch("/announcements/global/{announcement_id}", response_model=ToolAnnouncementInDB)
async def update_global_announcement(
    announcement_id: int,
    body: ToolAnnouncementUpdate,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_session),
):
    ensure_admin(current_user)
    tool_id = _resolve_global_announcement_tool_id(db)
    row = db.get(ToolAnnouncement, announcement_id)
    if not row or (row.visibility or "global") != "global":
        raise HTTPException(status_code=404, detail="公告不存在")
    if body.title is not None:
        row.title = body.title.strip()
    if body.content is not None:
        row.content = body.content.strip()
    if body.is_enabled is not None:
        row.is_enabled = body.is_enabled
    if body.start_at is not None:
        row.start_at = body.start_at
    if body.end_at is not None:
        row.end_at = body.end_at
    if row.end_at and row.start_at and row.end_at < row.start_at:
        raise HTTPException(status_code=400, detail="结束时间不能早于开始时间")
    if body.priority is not None:
        row.priority = _normalize_announcement_priority(body.priority)
    text_color, background_color = _priority_colors(row.priority or "notice")
    if body.scroll_speed_seconds is not None:
        row.scroll_speed_seconds = body.scroll_speed_seconds
    if body.font_family is not None:
        row.font_family = (body.font_family or "").strip() or None
    if body.font_size_px is not None:
        row.font_size_px = body.font_size_px
    row.text_color = text_color
    row.background_color = background_color
    row.visibility = "global"
    row.tool_id = tool_id
    row.disable_feature_slugs_json = "[]"
    row.updated_at = datetime.utcnow()
    db.add(row)
    db.commit()
    db.refresh(row)
    return _build_announcement_schema(row)


@router.get("/users/{user_id}/roles", response_model=UserRolesResponse)
async def get_user_roles(
    user_id: int,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_session),
):
    ensure_admin(current_user)
    user = db.get(User, user_id)
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在")

    roles = db.exec(
        select(Role.name).join(UserRole, UserRole.role_id == Role.id).where(UserRole.user_id == user_id)
    ).all()
    return UserRolesResponse(user_id=user_id, roles=roles)


@router.post("/users/{user_id}/approve", response_model=UserInDB)
async def approve_user_registration(
    user_id: int,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_session),
):
    """通过待审核用户的注册申请，通过后用户方可登录。"""
    ensure_admin(current_user)
    u = db.get(User, user_id)
    if not u:
        raise HTTPException(status_code=404, detail="用户不存在")
    if u.is_approved:
        raise HTTPException(status_code=400, detail="用户已审核通过")
    u.is_approved = True
    db.add(u)
    db.commit()
    db.refresh(u)
    return UserInDB.model_validate(u)


@router.post("/users/import-excel", response_model=AdminUserImportResponse)
async def import_users_from_excel(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_session),
):
    """管理员通过 Excel 批量创建普通用户（默认已审核通过，初始密码=邮箱）。"""
    ensure_admin(current_user)
    if not file.filename:
        raise HTTPException(status_code=400, detail="请上传 Excel 文件")
    filename = file.filename.lower()
    if not (filename.endswith(".xlsx") or filename.endswith(".xlsm")):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx / .xlsm 文件")

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="上传文件为空")

    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=400, detail=f"读取 Excel 失败：{exc}") from exc

    try:
        sheet = workbook.active
        row_iter = sheet.iter_rows(values_only=True)
        header = next(row_iter, None)
        if not header:
            raise HTTPException(status_code=400, detail="Excel 缺少表头")

        header_aliases = {
            "username": "username",
            "用户名": "username",
            "账号": "username",
            "user": "username",
            "email": "email",
            "邮箱": "email",
            "mail": "email",
            "fullname": "full_name",
            "name": "full_name",
            "姓名": "full_name",
            "department": "department",
            "dept": "department",
            "部门": "department",
        }
        col_index: dict[str, int] = {}
        for idx, raw in enumerate(header):
            key = header_aliases.get(_normalize_excel_header(raw))
            if key and key not in col_index:
                col_index[key] = idx
        if "email" not in col_index:
            raise HTTPException(status_code=400, detail="表头缺少邮箱列（email/邮箱）")

        tool_user_role = get_role_by_name(db, "tool_user")
        used_emails: set[str] = set()
        used_usernames: set[str] = set()
        created_users: list[UserInDB] = []
        skipped_items: list[AdminUserImportIssue] = []
        total_rows = 0

        for row_num, row in enumerate(row_iter, start=2):
            if not row:
                continue
            email_raw = row[col_index["email"]] if col_index["email"] < len(row) else None
            username_raw = (
                row[col_index["username"]]
                if "username" in col_index and col_index["username"] < len(row)
                else None
            )
            full_name_raw = (
                row[col_index["full_name"]]
                if "full_name" in col_index and col_index["full_name"] < len(row)
                else None
            )
            department_raw = (
                row[col_index["department"]]
                if "department" in col_index and col_index["department"] < len(row)
                else None
            )
            if all(str(item or "").strip() == "" for item in [email_raw, username_raw, full_name_raw, department_raw]):
                continue
            total_rows += 1

            try:
                email = _extract_email(email_raw)
            except ValueError as exc:
                skipped_items.append(
                    AdminUserImportIssue(row=row_num, email=str(email_raw or ""), reason=str(exc))
                )
                continue
            if email in used_emails:
                skipped_items.append(
                    AdminUserImportIssue(row=row_num, email=email, reason="文件内邮箱重复")
                )
                continue
            email_exists = db.exec(select(User.id).where(User.email == email)).first()
            if email_exists:
                skipped_items.append(
                    AdminUserImportIssue(row=row_num, email=email, reason="邮箱已存在")
                )
                continue

            username_seed = str(username_raw or "").strip()
            if not username_seed:
                username_seed = email.split("@", 1)[0]
            username = _pick_username(username_seed, used_usernames, db)
            full_name = str(full_name_raw or "").strip() or username
            department = str(department_raw or "").strip() or "未分配"

            row_user = User(
                username=username,
                email=email,
                hashed_password=get_password_hash(email),
                full_name=full_name,
                department=department,
                is_active=True,
                is_superuser=False,
                is_approved=True,
            )
            db.add(row_user)
            db.flush()
            db.add(UserRole(user_id=row_user.id, role_id=tool_user_role.id))
            used_emails.add(email)
            created_users.append(UserInDB.model_validate(row_user))

        db.commit()
        return AdminUserImportResponse(
            total_rows=total_rows,
            created_count=len(created_users),
            skipped_count=len(skipped_items),
            created_users=created_users,
            skipped_items=skipped_items,
        )
    except HTTPException:
        db.rollback()
        raise
    except Exception as exc:  # noqa: BLE001
        db.rollback()
        raise HTTPException(status_code=500, detail=f"批量导入失败：{exc}") from exc
    finally:
        workbook.close()


@router.get("/users/import-excel/template")
async def download_user_import_template(
    current_user: User = Depends(get_current_active_user),
):
    """下载批量导入普通用户的 Excel 模板。"""
    ensure_admin(current_user)
    wb = Workbook()
    ws = wb.active
    ws.title = "users"
    ws.append(["邮箱", "用户名", "姓名", "部门"])
    ws.append(["alice@example.com", "alice", "Alice", "研发部"])
    ws.append(["bob@example.com", "", "Bob", "测试部"])
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    output = io.BytesIO()
    wb.save(output)
    wb.close()
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="user-import-template.xlsx"'},
    )


@router.delete("/users/{user_id}", response_model=SuccessResponse)
async def admin_delete_user(
    user_id: int,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_session),
):
    """管理员注销其他用户账号（不可注销自己，需使用个人资料中的注销）。"""
    ensure_admin(current_user)
    if user_id == current_user.id:
        raise HTTPException(
            status_code=400,
            detail="不能在此删除自己，请前往个人资料页注销账号",
        )
    target = db.get(User, user_id)
    if not target:
        raise HTTPException(status_code=404, detail="用户不存在")
    if target.is_superuser:
        raise HTTPException(
            status_code=400,
            detail="超管账号不可通过管理接口注销，请先取消该用户的超管权限",
        )
    delete_user_and_related(db, user_id)
    return SuccessResponse(message="用户已删除")


@router.post("/users/{user_id}/roles", response_model=SuccessResponse)
async def assign_user_role(
    user_id: int,
    payload: RoleAssignmentRequest,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_session),
):
    ensure_admin(current_user)
    user = db.get(User, user_id)
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在")

    role = get_role_by_name(db, payload.role_name)
    exists = db.exec(
        select(UserRole).where(UserRole.user_id == user_id, UserRole.role_id == role.id)
    ).first()
    if exists:
        return SuccessResponse(message=f"角色「{payload.role_name}」已分配")

    db.add(UserRole(user_id=user_id, role_id=role.id))
    db.commit()
    return SuccessResponse(message=f"角色「{payload.role_name}」分配成功")


@router.delete("/users/{user_id}/roles/{role_name}", response_model=SuccessResponse)
async def revoke_user_role(
    user_id: int,
    role_name: str,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_session),
):
    ensure_admin(current_user)
    role = get_role_by_name(db, role_name)
    user_role = db.exec(
        select(UserRole).where(UserRole.user_id == user_id, UserRole.role_id == role.id)
    ).first()
    if not user_role:
        raise HTTPException(status_code=404, detail="未找到角色分配关系")

    db.delete(user_role)
    db.commit()
    return SuccessResponse(message=f"角色「{role_name}」撤销成功")


@router.get("/tools/{tool_id}/owners", response_model=List[ToolOwnerWithUser])
async def list_tool_owners(
    tool_id: int,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_session),
):
    ensure_admin(current_user)
    tool = db.get(Tool, tool_id)
    if not tool:
        raise HTTPException(status_code=404, detail="工具不存在")

    owners = db.exec(select(ToolOwner).where(ToolOwner.tool_id == tool_id)).all()
    result = []
    for owner in owners:
        user = db.get(User, owner.user_id)
        if user:
            result.append(ToolOwnerWithUser(**owner.dict(), user=user))
    return result


@router.post("/tools/{tool_id}/owners/{user_id}", response_model=SuccessResponse)
async def assign_tool_owner(
    tool_id: int,
    user_id: int,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_session),
):
    ensure_admin(current_user)
    tool = db.get(Tool, tool_id)
    if not tool:
        raise HTTPException(status_code=404, detail="工具不存在")
    user = db.get(User, user_id)
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在")

    exists = db.exec(
        select(ToolOwner).where(ToolOwner.tool_id == tool_id, ToolOwner.user_id == user_id)
    ).first()
    if exists:
        return SuccessResponse(message="工具负责人已分配")

    db.add(ToolOwner(tool_id=tool_id, user_id=user_id))

    role = get_role_by_name(db, "tool_owner")
    role_exists = db.exec(
        select(UserRole).where(UserRole.user_id == user_id, UserRole.role_id == role.id)
    ).first()
    if not role_exists:
        db.add(UserRole(user_id=user_id, role_id=role.id))

    db.commit()
    return SuccessResponse(message="工具负责人分配成功")


@router.delete("/tools/{tool_id}/owners/{user_id}", response_model=SuccessResponse)
async def remove_tool_owner(
    tool_id: int,
    user_id: int,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_session),
):
    ensure_admin(current_user)
    mapping = db.exec(
        select(ToolOwner).where(ToolOwner.tool_id == tool_id, ToolOwner.user_id == user_id)
    ).first()
    if not mapping:
        raise HTTPException(status_code=404, detail="未找到工具负责人分配关系")

    db.delete(mapping)
    db.commit()
    return SuccessResponse(message="工具负责人移除成功")


@router.get("/my-owner-tools", response_model=List[int])
async def get_my_owner_tools(
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_session),
):
    if current_user.is_superuser:
        return []
    tool_ids = db.exec(
        select(ToolOwner.tool_id).where(ToolOwner.user_id == current_user.id)
    ).all()
    return tool_ids


def _recipient_user_ids_for_tool(db: Session, tool_id: int) -> set[int]:
    """已获批用户 + 工具负责人（用于状态变更通知）"""
    s: set[int] = set()
    for p in db.exec(
        select(UserToolPermission).where(
            UserToolPermission.tool_id == tool_id,
            UserToolPermission.status == PermissionStatus.APPROVED,
        )
    ).all():
        s.add(p.user_id)
    for o in db.exec(select(ToolOwner).where(ToolOwner.tool_id == tool_id)).all():
        s.add(o.user_id)
    return s


@router.patch("/tools/{tool_id}/status", response_model=ToolInDB)
async def update_tool_status(
    tool_id: int,
    body: ToolStatusUpdate,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_session),
):
    tool = db.get(Tool, tool_id)
    if not tool:
        raise HTTPException(status_code=404, detail="工具不存在")
    ensure_permission_reviewer(db, current_user, tool_id)

    if tool.is_active == body.is_active:
        return tool

    tool.is_active = body.is_active
    db.add(tool)
    db.commit()
    db.refresh(tool)

    status_label = "可用" if tool.is_active else "暂不可用"
    for uid in _recipient_user_ids_for_tool(db, tool_id):
        db.add(
            Notification(
                user_id=uid,
                title=f"工具「{tool.name}」状态变更",
                message=f"工具「{tool.name}」当前状态为「{status_label}」。",
                notification_type="tool",
                related_id=tool_id,
            )
        )
    db.commit()
    return tool


@router.get("/tools/{tool_id}/usage-logs", response_model=PaginatedAPIAccessLogs)
async def get_tool_usage_logs(
    tool_id: int,
    skip: int = 0,
    limit: int = 100,
    username: str | None = None,
    q: str | None = None,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_session),
):
    tool = db.get(Tool, tool_id)
    if not tool:
        raise HTTPException(status_code=404, detail="工具不存在")
    ensure_permission_reviewer(db, current_user, tool_id)

    limit = min(max(limit, 1), 500)
    statement = (
        select(APIAccessLog)
        .where(APIAccessLog.tool_id == tool_id, APIAccessLog.feature_name != None)
    )
    if username and username.strip():
        statement = statement.where(
            APIAccessLog.username.ilike(f"%{username.strip()}%")
        )
    if q and q.strip():
        pattern = f"%{q.strip()}%"
        statement = statement.where(
            or_(
                APIAccessLog.path.ilike(pattern),
                APIAccessLog.feature_name.ilike(pattern),
                APIAccessLog.method.ilike(pattern),
                APIAccessLog.behavior_label.ilike(pattern),
            )
        )
    count_stmt = select(func.count(APIAccessLog.id)).where(
        APIAccessLog.tool_id == tool_id,
        APIAccessLog.feature_name != None,
    )
    if username and username.strip():
        count_stmt = count_stmt.where(
            APIAccessLog.username.ilike(f"%{username.strip()}%")
        )
    if q and q.strip():
        pattern = f"%{q.strip()}%"
        count_stmt = count_stmt.where(
            or_(
                APIAccessLog.path.ilike(pattern),
                APIAccessLog.feature_name.ilike(pattern),
                APIAccessLog.method.ilike(pattern),
                APIAccessLog.behavior_label.ilike(pattern),
            )
        )
    raw_total = db.exec(count_stmt).first()
    if raw_total is None:
        total = 0
    elif isinstance(raw_total, int):
        total = raw_total
    else:
        total = int(raw_total[0])

    logs = db.exec(
        statement.order_by(APIAccessLog.created_at.desc()).offset(skip).limit(limit)
    ).all()
    return PaginatedAPIAccessLogs(
        total=total,
        items=[build_access_log_item(db, log) for log in logs],
    )


@router.get("/tools/{tool_id}/license-users", response_model=PaginatedToolLicenseUsers)
async def list_tool_license_users(
    tool_id: int,
    skip: int = 0,
    limit: int = 20,
    search: str | None = None,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_session),
):
    tool = db.get(Tool, tool_id)
    if not tool:
        raise HTTPException(status_code=404, detail="工具不存在")
    ensure_permission_reviewer(db, current_user, tool_id)

    perms = db.exec(
        select(UserToolPermission).where(
            UserToolPermission.tool_id == tool_id,
            UserToolPermission.status == PermissionStatus.APPROVED,
        )
    ).all()

    last_by_user: dict[int, datetime] = {}
    agg_rows = db.exec(
        select(APIAccessLog.user_id, func.max(APIAccessLog.created_at))
        .where(
            APIAccessLog.tool_id == tool_id,
            APIAccessLog.feature_name != None,
            APIAccessLog.user_id != None,
        )
        .group_by(APIAccessLog.user_id)
    ).all()
    for row in agg_rows:
        cells = tuple(row)
        if len(cells) < 2:
            continue
        uid, ts = cells[0], cells[1]
        if uid is not None and ts is not None:
            last_by_user[int(uid)] = ts

    result: List[ToolLicenseUserRow] = []
    for p in perms:
        user = db.get(User, p.user_id)
        if not user:
            continue
        granted_at = p.reviewed_at or p.applied_at
        result.append(
            ToolLicenseUserRow(
                user=user,
                granted_at=granted_at,
                expires_at=p.expires_at,
                last_used_at=last_by_user.get(p.user_id),
            )
        )

    limit = min(max(limit, 1), 500)
    result.sort(key=lambda r: r.user.username.lower())
    if search and search.strip():
        q = search.strip().lower()
        result = [
            r
            for r in result
            if q in r.user.username.lower()
            or q in (r.user.email or "").lower()
            or (r.user.full_name and q in r.user.full_name.lower())
        ]
    total = len(result)
    return PaginatedToolLicenseUsers(
        total=total,
        items=result[skip: skip + limit],
    )


@router.delete("/tools/{tool_id}/license-users/{user_id}", response_model=SuccessResponse)
async def revoke_tool_user_license(
    tool_id: int,
    user_id: int,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_session),
):
    tool = db.get(Tool, tool_id)
    if not tool:
        raise HTTPException(status_code=404, detail="工具不存在")
    ensure_permission_reviewer(db, current_user, tool_id)

    perm = db.exec(
        select(UserToolPermission).where(
            UserToolPermission.tool_id == tool_id,
            UserToolPermission.user_id == user_id,
            UserToolPermission.status == PermissionStatus.APPROVED,
        )
    ).first()
    if not perm:
        raise HTTPException(
            status_code=404,
            detail="未找到该用户在该工具下的已批准权限",
        )

    db.delete(perm)
    db.commit()

    db.add(
        Notification(
            user_id=user_id,
            title=f"工具「{tool.name}」使用权限已取消",
            message=f"管理员或工具负责人已取消您对「{tool.name}」的使用权限。",
            notification_type="permission",
            related_id=tool_id,
        )
    )
    db.commit()
    return SuccessResponse(message="权限已撤销")


@router.get("/tools/{tool_id}/feedback", response_model=PaginatedFeedbackWithUser)
async def list_tool_feedback(
    tool_id: int,
    skip: int = 0,
    limit: int = 20,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_session),
):
    tool = db.get(Tool, tool_id)
    if not tool:
        raise HTTPException(status_code=404, detail="工具不存在")
    ensure_permission_reviewer(db, current_user, tool_id)

    limit = min(max(limit, 1), 500)
    raw_total = db.exec(
        select(func.count(Feedback.id)).where(
            Feedback.tool_id == tool_id,
            Feedback.category == FeedbackCategory.TOOL_USAGE.value,
        )
    ).first()
    if raw_total is None:
        total = 0
    elif isinstance(raw_total, int):
        total = raw_total
    else:
        total = int(raw_total[0])

    rows = db.exec(
        select(Feedback)
        .where(
            Feedback.tool_id == tool_id,
            Feedback.category == FeedbackCategory.TOOL_USAGE.value,
        )
        .order_by(Feedback.created_at.desc())
        .offset(skip)
        .limit(limit)
    ).all()
    return PaginatedFeedbackWithUser(
        total=total,
        items=[build_feedback_with_user(db, f) for f in rows],
    )


@router.get("/feedback", response_model=PaginatedFeedbackWithUser)
async def list_global_feedback(
    category: Literal["system_feedback", "new_tool_suggestion"],
    skip: int = 0,
    limit: int = 100,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_session),
):
    ensure_admin(current_user)
    limit = min(max(limit, 1), 500)
    raw_total = db.exec(
        select(func.count(Feedback.id)).where(Feedback.category == category)
    ).first()
    if raw_total is None:
        total = 0
    elif isinstance(raw_total, int):
        total = raw_total
    else:
        total = int(raw_total[0])

    rows = db.exec(
        select(Feedback)
        .where(Feedback.category == category)
        .order_by(Feedback.created_at.desc())
        .offset(skip)
        .limit(limit)
    ).all()
    return PaginatedFeedbackWithUser(
        total=total,
        items=[build_feedback_with_user(db, f) for f in rows],
    )


@router.get("/feedback/counts", response_model=FeedbackCountsResponse)
async def feedback_counts(
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_session),
):
    ensure_admin(current_user)
    sf = db.exec(
        select(Feedback).where(Feedback.category == FeedbackCategory.SYSTEM_FEEDBACK.value)
    ).all()
    ns = db.exec(
        select(Feedback).where(Feedback.category == FeedbackCategory.NEW_TOOL_SUGGESTION.value)
    ).all()
    n_sf = len(sf)
    n_ns = len(ns)
    return FeedbackCountsResponse(
        system_feedback=n_sf,
        new_tool_suggestion=n_ns,
        total=n_sf + n_ns,
    )


def _audit_log_filter_conditions(
    user_id: int | None,
    tool_id: int | None,
    username: str | None,
    q: str | None = None,
):
    conditions = []
    if user_id is not None:
        conditions.append(APIAccessLog.user_id == user_id)
    if tool_id is not None:
        conditions.append(APIAccessLog.tool_id == tool_id)
    if username and username.strip():
        pattern = f"%{username.strip()}%"
        conditions.append(APIAccessLog.username.ilike(pattern))
    if q and q.strip():
        pattern = f"%{q.strip()}%"
        conditions.append(
            or_(
                APIAccessLog.path.ilike(pattern),
                APIAccessLog.feature_name.ilike(pattern),
                APIAccessLog.method.ilike(pattern),
                APIAccessLog.behavior_label.ilike(pattern),
            )
        )
    return conditions


@router.get("/audit-logs/export")
async def export_audit_logs_csv(
    user_id: int | None = None,
    tool_id: int | None = None,
    username: str | None = None,
    q: str | None = None,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_session),
):
    ensure_admin(current_user)
    max_rows = 10_000
    statement = select(APIAccessLog)
    for cond in _audit_log_filter_conditions(user_id, tool_id, username, q):
        statement = statement.where(cond)
    logs = db.exec(
        statement.order_by(APIAccessLog.created_at.desc()).limit(max_rows)
    ).all()

    tool_ids = {log.tool_id for log in logs if log.tool_id}
    tools_by_id: dict[int, Tool] = {}
    if tool_ids:
        for t in db.exec(select(Tool).where(Tool.id.in_(list(tool_ids)))).all():
            tools_by_id[t.id] = t

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(
        [
            "id",
            "created_at",
            "username",
            "user_id",
            "method",
            "path",
            "feature_name",
            "behavior_label",
            "tool_id",
            "status_code",
            "latency_ms",
            "client_ip",
            "query_string",
        ]
    )
    for log in logs:
        bl = log.behavior_label
        if not bl and log.tool_id and log.feature_name:
            tool = tools_by_id.get(log.tool_id)
            bl = resolve_behavior_label_from_tool(tool, log.feature_name)
        writer.writerow(
            [
                log.id,
                _format_ts_cst8(log.created_at),
                log.username or "",
                log.user_id if log.user_id is not None else "",
                log.method,
                log.path,
                log.feature_name or "",
                bl or "",
                log.tool_id if log.tool_id is not None else "",
                log.status_code,
                log.latency_ms,
                log.client_ip or "",
                log.query_string or "",
            ]
        )
    payload = "\ufeff" + buf.getvalue()
    return StreamingResponse(
        iter([payload.encode("utf-8")]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": 'attachment; filename="audit-logs.csv"'},
    )


@router.get("/audit-logs", response_model=PaginatedAPIAccessLogs)
async def get_all_audit_logs(
    skip: int = 0,
    limit: int = 20,
    user_id: int | None = None,
    tool_id: int | None = None,
    username: str | None = None,
    q: str | None = None,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_session),
):
    ensure_admin(current_user)
    limit = min(max(limit, 1), 500)

    count_stmt = select(func.count(APIAccessLog.id))
    statement = select(APIAccessLog)
    for cond in _audit_log_filter_conditions(user_id, tool_id, username, q):
        count_stmt = count_stmt.where(cond)
        statement = statement.where(cond)

    raw_total = db.exec(count_stmt).first()
    if raw_total is None:
        total = 0
    elif isinstance(raw_total, int):
        total = raw_total
    else:
        total = int(raw_total[0])

    logs = db.exec(
        statement.order_by(APIAccessLog.created_at.desc()).offset(skip).limit(limit)
    ).all()
    items = [build_access_log_item(db, log) for log in logs]
    return PaginatedAPIAccessLogs(total=total, items=items)

@router.get("/permissions/pending", response_model=List[PermissionWithDetails])
async def get_pending_permissions(
    skip: int = 0,
    limit: int = 100,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_session),
):
    if current_user.is_superuser:
        statement = select(UserToolPermission).where(
            UserToolPermission.status == PermissionStatus.PENDING
        ).offset(skip).limit(limit)
    else:
        owner_assignments = db.exec(
            select(ToolOwner.tool_id).where(ToolOwner.user_id == current_user.id)
        ).all()
        if not owner_assignments:
            raise HTTPException(status_code=403, detail="仅工具负责人可审核权限")
        statement = select(UserToolPermission).where(
            UserToolPermission.status == PermissionStatus.PENDING,
            UserToolPermission.tool_id.in_(owner_assignments),
        ).offset(skip).limit(limit)

    permissions = db.exec(statement).all()

    # 加载关联数据
    result = []
    for perm in permissions:
        user = db.get(User, perm.user_id)
        tool = db.get(Tool, perm.tool_id)

        if user and tool:
            reviewer = db.get(User, perm.reviewed_by) if perm.reviewed_by else None
            perm_with_details = PermissionWithDetails(
                **perm.dict(),
                user=user,
                tool=tool,
                reviewer=reviewer
            )
            result.append(perm_with_details)

    return result

@router.post("/permissions/{permission_id}/approve", response_model=SuccessResponse)
async def approve_permission(
    permission_id: int,
    update_data: PermissionUpdate,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_session),
):
    permission = db.get(UserToolPermission, permission_id)
    if not permission:
        raise HTTPException(status_code=404, detail="权限申请不存在")

    ensure_permission_reviewer(db, current_user, permission.tool_id)

    if permission.status != PermissionStatus.PENDING:
        raise HTTPException(
            status_code=400,
            detail="该权限申请不处于待审核状态"
        )

    # 更新权限状态
    permission.status = PermissionStatus.APPROVED
    permission.reviewed_by = current_user.id
    permission.reviewed_at = datetime.utcnow()
    if update_data.review_notes:
        permission.review_notes = update_data.review_notes
    if update_data.expires_at:
        permission.expires_at = update_data.expires_at

    db.add(permission)
    db.commit()

    tool = db.get(Tool, permission.tool_id)
    tool_name = tool.name if tool else f"ID {permission.tool_id}"

    # 创建通知（审核结果 + 后续使用引导）
    notification = Notification(
        user_id=permission.user_id,
        title="权限申请已批准",
        message=f"您对工具「{tool_name}」的权限申请已被批准。",
        notification_type="permission",
        related_id=permission_id
    )
    db.add(notification)
    notification_tool = Notification(
        user_id=permission.user_id,
        title=f"工具「{tool_name}」已就绪",
        message=f"请在「我的工具」或「所有工具」中打开该工具开始使用。",
        notification_type="tool",
        related_id=permission.tool_id,
    )
    db.add(notification_tool)
    db.commit()

    return SuccessResponse(message="权限申请已批准")

@router.post("/permissions/{permission_id}/reject", response_model=SuccessResponse)
async def reject_permission(
    permission_id: int,
    update_data: PermissionUpdate,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_session),
):
    if not update_data.review_notes:
        raise HTTPException(
            status_code=400,
            detail="拒绝时必须填写审核说明"
        )

    permission = db.get(UserToolPermission, permission_id)
    if not permission:
        raise HTTPException(status_code=404, detail="权限申请不存在")

    ensure_permission_reviewer(db, current_user, permission.tool_id)

    if permission.status != PermissionStatus.PENDING:
        raise HTTPException(
            status_code=400,
            detail="该权限申请不处于待审核状态"
        )

    # 更新权限状态
    permission.status = PermissionStatus.REJECTED
    permission.reviewed_by = current_user.id
    permission.reviewed_at = datetime.utcnow()
    permission.review_notes = update_data.review_notes

    db.add(permission)
    db.commit()

    tool = db.get(Tool, permission.tool_id)
    tool_name = tool.name if tool else f"ID {permission.tool_id}"

    # 创建拒绝通知
    notification = Notification(
        user_id=permission.user_id,
        title="权限申请被拒绝",
        message=f"您对工具「{tool_name}」的权限申请已被拒绝。原因：{update_data.review_notes}",
        notification_type="permission",
        related_id=permission_id
    )
    db.add(notification)
    db.commit()

    return SuccessResponse(message="权限申请已拒绝")


def _notification_message_for_release(tool_name: str, row: ToolRelease) -> str:
    header = f"版本：{row.version}"
    if row.spec_revision:
        header += f"　规格修订：{row.spec_revision}"
    body = f"{row.title}\n\n{row.changelog}"
    if len(body) > 2500:
        body = body[:2500] + "\n…（完整说明请在工具的「更新记录」中查看）"
    return f"{header}\n\n{body}"


@router.post("/tools/{tool_id}/releases", response_model=ToolReleaseInDB)
async def publish_tool_release(
    tool_id: int,
    body: ToolReleasePublish,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_session),
):
    tool = db.get(Tool, tool_id)
    if not tool:
        raise HTTPException(status_code=404, detail="工具不存在")
    ensure_permission_reviewer(db, current_user, tool_id)

    now = datetime.utcnow()
    sv = body.version.strip()
    sr = body.spec_revision.strip() if body.spec_revision else None
    row = ToolRelease(
        tool_id=tool_id,
        version=sv,
        spec_revision=sr,
        title=body.title.strip(),
        changelog=body.changelog.strip(),
        published_at=now,
        published_by=current_user.id,
    )
    tool.version = sv
    if sr:
        tool.spec_revision = sr
    db.add(row)
    db.add(tool)
    db.commit()
    db.refresh(row)

    if body.notify_users:
        title = f"工具「{tool.name}」已发版 {row.version}"
        msg = _notification_message_for_release(tool.name, row)
        for uid in _recipient_user_ids_for_tool(db, tool_id):
            db.add(
                Notification(
                    user_id=uid,
                    title=title,
                    message=msg,
                    notification_type="tool_release",
                    related_id=tool_id,
                )
            )
        db.commit()

    return ToolReleaseInDB.model_validate(row)
