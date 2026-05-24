import io
import re
from datetime import datetime
from typing import List

from email_validator import EmailNotValidError, validate_email
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlmodel import Session, select

from app.api.v1.admin_common import get_role_by_name
from app.api.v1.auth import get_password_hash
from app.api.v1.rbac import ensure_platform_staff, ensure_super_admin, has_role
from app.api.v1.users import get_current_active_user
from app.database import get_session
from app.models import (
    Notification,
    PermissionStatus,
    Role,
    Tool,
    ToolDisplayConfig,
    User,
    UserRole,
    UserToolPermission,
)
from app.schemas import (
    AdminResetPasswordRequest,
    AdminToolAssignmentOption,
    AdminUserAllowedToolsResponse,
    AdminUserAllowedToolsUpdate,
    AdminUserImportIssue,
    AdminUserImportResponse,
    RoleAssignmentRequest,
    SuccessResponse,
    UserInDB,
    UserRolesResponse,
)
from app.services.user_deletion import delete_user_and_related

router = APIRouter()

ADMIN_TOOL_GRANT_REASON = "管理员分配可用工具"


def _normalize_excel_header(value: object) -> str:
    text = str(value or "").strip().lower()
    return re.sub(r"[\s_\-]+", "", text)


def _parse_excel_tools_cell(db: Session, raw: object) -> list[int]:
    text = str(raw or "").strip()
    if not text:
        return []
    parts = [p.strip() for p in re.split(r"[,，;；|、\s]+", text) if p.strip()]
    if not parts:
        return []
    out: list[int] = []
    seen: set[int] = set()
    for part in parts:
        tid: int | None = None
        if part.isdigit():
            tid = int(part)
            tool = db.get(Tool, tid)
            if not tool or not tool.is_active:
                raise ValueError(f"工具 ID {tid} 不存在或已停用")
        else:
            tool = db.exec(select(Tool).where(Tool.name == part)).first()
            if not tool or not tool.is_active:
                raise ValueError(f"未知或已停用的工具标识：{part}")
            tid = int(tool.id)
        if tid not in seen:
            seen.add(tid)
            out.append(tid)
    return out


def _apply_admin_user_allowed_tools(
    db: Session,
    user_id: int,
    tool_ids: list[int],
    reviewed_by_user_id: int,
) -> None:
    want: set[int] = set()
    for raw in tool_ids:
        try:
            tid = int(raw)
        except (TypeError, ValueError) as exc:
            raise HTTPException(status_code=400, detail=f"非法的工具 ID：{raw}") from exc
        if tid <= 0:
            raise HTTPException(status_code=400, detail=f"非法的工具 ID：{raw}")
        want.add(tid)
    for tid in want:
        tool = db.get(Tool, tid)
        if not tool:
            raise HTTPException(status_code=400, detail=f"工具不存在：{tid}")
        if not tool.is_active:
            raise HTTPException(status_code=400, detail=f"工具已停用，不可分配：{tool.name}")

    existing = db.exec(
        select(UserToolPermission).where(UserToolPermission.user_id == user_id)
    ).all()
    now = datetime.utcnow()
    for perm in existing:
        if perm.tool_id not in want:
            db.delete(perm)
    for tid in want:
        perm = db.exec(
            select(UserToolPermission).where(
                UserToolPermission.user_id == user_id,
                UserToolPermission.tool_id == tid,
            )
        ).first()
        if perm:
            perm.status = PermissionStatus.APPROVED
            perm.applied_reason = ADMIN_TOOL_GRANT_REASON
            perm.reviewed_by = reviewed_by_user_id
            perm.reviewed_at = now
            perm.review_notes = None
            perm.expires_at = None
            db.add(perm)
        else:
            db.add(
                UserToolPermission(
                    user_id=user_id,
                    tool_id=tid,
                    status=PermissionStatus.APPROVED,
                    applied_reason=ADMIN_TOOL_GRANT_REASON,
                    reviewed_by=reviewed_by_user_id,
                    reviewed_at=now,
                )
            )


def _sanitize_username(raw: str) -> str:
    text = re.sub(r"[^a-zA-Z0-9_.-]+", "-", raw.strip().lower()).strip("-._")
    return text[:50] if text else "user"


def _pick_username(base: str, used_usernames: set[str], db: Session) -> str:
    base_candidate = _sanitize_username(base)
    candidate = base_candidate
    suffix = 1
    while True:
        if candidate not in used_usernames:
            exists = db.exec(select(User.id).where(User.username == candidate)).first()
            if not exists:
                used_usernames.add(candidate)
                return candidate
        suffix += 1
        clipped = base_candidate[:40] if len(base_candidate) > 40 else base_candidate
        candidate = f"{clipped}-{suffix}"


def _extract_email(raw: object) -> str:
    email = str(raw or "").strip().lower()
    if not email:
        raise ValueError("邮箱为空")
    try:
        normalized = validate_email(email, check_deliverability=False)
    except EmailNotValidError as exc:
        raise ValueError(f"邮箱格式非法：{exc}") from exc
    return normalized.normalized


@router.get("/users/{user_id}/roles", response_model=UserRolesResponse)
async def get_user_roles(
    user_id: int,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_session),
):
    ensure_platform_staff(db, current_user)
    user = db.get(User, user_id)
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在")
    roles = db.exec(
        select(Role.name).join(UserRole, UserRole.role_id == Role.id).where(UserRole.user_id == user_id)
    ).all()
    return UserRolesResponse(user_id=user_id, roles=roles)


@router.post("/users/{user_id}/approve", response_model=UserInDB)
async def approve_user_registration(
    user_id: int,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_session),
):
    ensure_platform_staff(db, current_user)
    user = db.get(User, user_id)
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在")
    if user.is_approved:
        raise HTTPException(status_code=400, detail="用户已审核通过")
    user.is_approved = True
    db.add(user)
    db.commit()
    db.refresh(user)
    db.add(
        Notification(
            user_id=user.id,
            title="注册审核已通过",
            message="您的账号已通过审核，现在可以正常登录并使用系统。",
            notification_type="system",
            related_id=user.id,
        )
    )
    db.commit()
    db.refresh(user)
    pa = has_role(db, user.id, "platform_admin")
    return UserInDB.model_validate(user).model_copy(update={"is_platform_admin": bool(pa)})


@router.post("/users/{user_id}/transfer-super-admin", response_model=UserInDB)
async def transfer_super_admin(
    user_id: int,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_session),
):
    ensure_super_admin(current_user)
    if user_id == current_user.id:
        raise HTTPException(status_code=400, detail="请选择要接任超级管理员的目标用户")
    target = db.get(User, user_id)
    if not target or not target.is_active:
        raise HTTPException(status_code=404, detail="目标用户不存在或已停用")
    if target.is_superuser:
        raise HTTPException(status_code=400, detail="该用户已是超级管理员")

    pa_role = db.exec(select(Role).where(Role.name == "platform_admin")).first()
    if pa_role:
        cur_has = db.exec(
            select(UserRole).where(
                UserRole.user_id == current_user.id,
                UserRole.role_id == pa_role.id,
            )
        ).first()
        if not cur_has:
            db.add(UserRole(user_id=current_user.id, role_id=pa_role.id))

    current_user.is_superuser = False
    target.is_superuser = True
    db.add(current_user)
    db.add(target)
    db.commit()
    db.refresh(target)
    db.add(
        Notification(
            user_id=target.id,
            title="您已被设为超级管理员",
            message="前任超级管理员已将身份转移给您，请重新登录以刷新权限。系统级配置仅超级管理员可操作。",
            notification_type="system",
            related_id=target.id,
        )
    )
    db.add(
        Notification(
            user_id=current_user.id,
            title="超级管理员身份已转移",
            message=f"已将超级管理员权限转移至用户「{target.username}」，您现为平台管理员。",
            notification_type="system",
            related_id=current_user.id,
        )
    )
    db.commit()
    pa = has_role(db, target.id, "platform_admin")
    return UserInDB.model_validate(target).model_copy(update={"is_platform_admin": bool(pa)})


def _list_tool_assignment_options_impl(
    current_user: User,
    db: Session,
) -> List[AdminToolAssignmentOption]:
    ensure_platform_staff(db, current_user)
    rows = db.exec(
        select(Tool, ToolDisplayConfig)
        .outerjoin(ToolDisplayConfig, ToolDisplayConfig.tool_id == Tool.id)
        .order_by(Tool.id)
    ).all()
    return [
        AdminToolAssignmentOption(
            id=tool.id,
            name=tool.name,
            display_name=(cfg.display_name if cfg else None),
            is_active=tool.is_active,
            runtime_status=tool.runtime_status,
        )
        for tool, cfg in rows
    ]


@router.get("/tool_assignment/options", response_model=List[AdminToolAssignmentOption])
async def list_tool_assignment_options(
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_session),
):
    return _list_tool_assignment_options_impl(current_user, db)


@router.get("/tool-assignment/options", response_model=List[AdminToolAssignmentOption], include_in_schema=False)
async def list_tool_assignment_options_hyphen_alias(
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_session),
):
    return _list_tool_assignment_options_impl(current_user, db)


def _get_user_allowed_tools_impl(
    user_id: int,
    current_user: User,
    db: Session,
) -> AdminUserAllowedToolsResponse:
    ensure_platform_staff(db, current_user)
    if not db.get(User, user_id):
        raise HTTPException(status_code=404, detail="用户不存在")
    perms = db.exec(
        select(UserToolPermission).where(
            UserToolPermission.user_id == user_id,
            UserToolPermission.status == PermissionStatus.APPROVED,
        )
    ).all()
    tool_ids = sorted({int(p.tool_id) for p in perms if p.tool_id is not None})
    return AdminUserAllowedToolsResponse(tool_ids=tool_ids)


@router.get("/users/{user_id}/allowed_tools", response_model=AdminUserAllowedToolsResponse)
async def get_user_allowed_tools(
    user_id: int,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_session),
):
    return _get_user_allowed_tools_impl(user_id, current_user, db)


@router.get("/users/{user_id}/allowed-tools", response_model=AdminUserAllowedToolsResponse, include_in_schema=False)
async def get_user_allowed_tools_hyphen_alias(
    user_id: int,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_session),
):
    return _get_user_allowed_tools_impl(user_id, current_user, db)


@router.put("/users/{user_id}/allowed_tools", response_model=AdminUserAllowedToolsResponse)
async def put_user_allowed_tools(
    user_id: int,
    body: AdminUserAllowedToolsUpdate,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_session),
):
    ensure_platform_staff(db, current_user)
    if not db.get(User, user_id):
        raise HTTPException(status_code=404, detail="用户不存在")
    _apply_admin_user_allowed_tools(db, user_id, body.tool_ids, current_user.id)
    db.commit()
    return _get_user_allowed_tools_impl(user_id, current_user, db)


@router.put("/users/{user_id}/allowed-tools", response_model=AdminUserAllowedToolsResponse, include_in_schema=False)
async def put_user_allowed_tools_hyphen_alias(
    user_id: int,
    body: AdminUserAllowedToolsUpdate,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_session),
):
    return await put_user_allowed_tools(user_id, body, current_user, db)


@router.post("/users/{user_id}/reset-password", response_model=SuccessResponse)
async def reset_user_password_by_admin(
    user_id: int,
    payload: AdminResetPasswordRequest,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_session),
):
    ensure_platform_staff(db, current_user)
    user = db.get(User, user_id)
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在")
    if not user.is_active:
        raise HTTPException(status_code=400, detail="用户已禁用，无法重置密码")

    user.hashed_password = get_password_hash(payload.new_password)
    user.updated_at = datetime.utcnow()
    db.add(user)
    db.add(
        Notification(
            user_id=user.id,
            title="账号密码已被管理员重置",
            message="您的账号密码已被管理员重置，请联系管理员获取新密码并尽快在个人资料页修改。",
            notification_type="system",
            related_id=user.id,
        )
    )
    db.commit()
    return SuccessResponse(message=f"用户「{user.username}」密码已重置")


@router.post("/users/import-excel", response_model=AdminUserImportResponse)
async def import_users_from_excel(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_session),
):
    ensure_platform_staff(db, current_user)
    if not file.filename:
        raise HTTPException(status_code=400, detail="请上传 Excel 文件")
    filename = file.filename.lower()
    if not (filename.endswith(".xlsx") or filename.endswith(".xlsm")):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx / .xlsm 文件")

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="上传文件为空")

    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=400, detail=f"读取 Excel 失败：{exc}") from exc

    try:
        sheet = workbook.active
        row_iter = sheet.iter_rows(values_only=True)
        header = next(row_iter, None)
        if not header:
            raise HTTPException(status_code=400, detail="Excel 缺少表头")

        header_aliases = {
            "username": "username",
            "用户名": "username",
            "账号": "username",
            "user": "username",
            "email": "email",
            "邮箱": "email",
            "mail": "email",
            "fullname": "full_name",
            "name": "full_name",
            "姓名": "full_name",
            "department": "department",
            "dept": "department",
            "部门": "department",
            "tools": "tools",
            "工具": "tools",
            "可用工具": "tools",
            "允许的工具": "tools",
            "toolids": "tools",
            "tool_ids": "tools",
        }
        col_index: dict[str, int] = {}
        for idx, raw in enumerate(header):
            key = header_aliases.get(_normalize_excel_header(raw))
            if key and key not in col_index:
                col_index[key] = idx
        if "email" not in col_index:
            raise HTTPException(status_code=400, detail="表头缺少邮箱列（email/邮箱）")

        tool_user_role = get_role_by_name(db, "tool_user")
        used_emails: set[str] = set()
        used_usernames: set[str] = set()
        created_users: list[UserInDB] = []
        skipped_items: list[AdminUserImportIssue] = []
        total_rows = 0

        for row_num, row in enumerate(row_iter, start=2):
            if not row:
                continue
            email_raw = row[col_index["email"]] if col_index["email"] < len(row) else None
            username_raw = row[col_index["username"]] if "username" in col_index and col_index["username"] < len(row) else None
            full_name_raw = row[col_index["full_name"]] if "full_name" in col_index and col_index["full_name"] < len(row) else None
            department_raw = row[col_index["department"]] if "department" in col_index and col_index["department"] < len(row) else None
            if all(str(item or "").strip() == "" for item in [email_raw, username_raw, full_name_raw, department_raw]):
                continue
            total_rows += 1

            try:
                email = _extract_email(email_raw)
            except ValueError as exc:
                skipped_items.append(AdminUserImportIssue(row=row_num, email=str(email_raw or ""), reason=str(exc)))
                continue
            if email in used_emails:
                skipped_items.append(AdminUserImportIssue(row=row_num, email=email, reason="文件内邮箱重复"))
                continue
            email_exists = db.exec(select(User.id).where(User.email == email)).first()
            if email_exists:
                skipped_items.append(AdminUserImportIssue(row=row_num, email=email, reason="邮箱已存在"))
                continue

            row_tool_ids: list[int] = []
            if "tools" in col_index:
                tools_raw = row[col_index["tools"]] if col_index["tools"] < len(row) else None
                try:
                    row_tool_ids = _parse_excel_tools_cell(db, tools_raw)
                except ValueError as exc:
                    skipped_items.append(AdminUserImportIssue(row=row_num, email=email, reason=str(exc)))
                    continue

            username_seed = str(username_raw or "").strip()
            if not username_seed:
                username_seed = email.split("@", 1)[0]
            username = _pick_username(username_seed, used_usernames, db)
            full_name = str(full_name_raw or "").strip() or username
            department = str(department_raw or "").strip() or "未分配"

            row_user = User(
                username=username,
                email=email,
                hashed_password=get_password_hash(email),
                full_name=full_name,
                department=department,
                is_active=True,
                is_superuser=False,
                is_approved=True,
            )
            db.add(row_user)
            db.flush()
            db.add(UserRole(user_id=row_user.id, role_id=tool_user_role.id))
            if row_tool_ids:
                _apply_admin_user_allowed_tools(db, int(row_user.id), row_tool_ids, current_user.id)
            used_emails.add(email)
            created_users.append(UserInDB.model_validate(row_user))

        db.commit()
        return AdminUserImportResponse(
            total_rows=total_rows,
            created_count=len(created_users),
            skipped_count=len(skipped_items),
            created_users=created_users,
            skipped_items=skipped_items,
        )
    except HTTPException:
        db.rollback()
        raise
    except Exception as exc:  # noqa: BLE001
        db.rollback()
        raise HTTPException(status_code=500, detail=f"批量导入失败：{exc}") from exc
    finally:
        workbook.close()


@router.get("/users/import-excel/template")
async def download_user_import_template(
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_session),
):
    _ = db
    ensure_platform_staff(db, current_user)
    wb = Workbook()
    ws = wb.active
    ws.title = "users"
    ws.append(["邮箱", "用户名", "姓名", "部门", "可用工具"])
    ws.append(["alice@example.com", "alice", "Alice", "研发部", "mos-integration-toolbox"])
    ws.append(["bob@example.com", "", "Bob", "测试部", ""])
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 40
    output = io.BytesIO()
    wb.save(output)
    wb.close()
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="user-import-template.xlsx"'},
    )


@router.delete("/users/{user_id}", response_model=SuccessResponse)
async def admin_delete_user(
    user_id: int,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_session),
):
    ensure_platform_staff(db, current_user)
    if user_id == current_user.id:
        raise HTTPException(status_code=400, detail="不能在此删除自己，请前往个人资料页注销账号")
    target = db.get(User, user_id)
    if not target:
        raise HTTPException(status_code=404, detail="用户不存在")
    if target.is_superuser:
        raise HTTPException(status_code=400, detail="超管账号不可通过管理接口注销，请先取消该用户的超管权限")
    delete_user_and_related(db, user_id)
    return SuccessResponse(message="用户已删除")


@router.post("/users/{user_id}/roles", response_model=SuccessResponse)
async def assign_user_role(
    user_id: int,
    payload: RoleAssignmentRequest,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_session),
):
    ensure_platform_staff(db, current_user)
    if payload.role_name == "platform_admin":
        ensure_super_admin(current_user)
    user = db.get(User, user_id)
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在")
    if payload.role_name == "platform_admin" and user.is_superuser:
        raise HTTPException(status_code=400, detail="超级管理员无需再分配「管理员」角色")

    role = get_role_by_name(db, payload.role_name)
    exists = db.exec(select(UserRole).where(UserRole.user_id == user_id, UserRole.role_id == role.id)).first()
    if exists:
        return SuccessResponse(message=f"角色「{payload.role_name}」已分配")

    db.add(UserRole(user_id=user_id, role_id=role.id))
    db.commit()
    return SuccessResponse(message=f"角色「{payload.role_name}」分配成功")


@router.delete("/users/{user_id}/roles/{role_name}", response_model=SuccessResponse)
async def revoke_user_role(
    user_id: int,
    role_name: str,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_session),
):
    ensure_platform_staff(db, current_user)
    if role_name == "platform_admin":
        ensure_super_admin(current_user)
    role = get_role_by_name(db, role_name)
    user_role = db.exec(select(UserRole).where(UserRole.user_id == user_id, UserRole.role_id == role.id)).first()
    if not user_role:
        raise HTTPException(status_code=404, detail="未找到角色分配关系")

    db.delete(user_role)
    db.commit()
    return SuccessResponse(message=f"角色「{role_name}」撤销成功")
